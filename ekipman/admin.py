from django.contrib import admin
from django.http import HttpResponse
from .models import Ekipman, QRSorguKaydi, Gorev  # Gorev modelini ekledik
import openpyxl
from openpyxl.utils import get_column_letter

@admin.register(Ekipman)
class EkipmanAdmin(admin.ModelAdmin):
    list_display = ('adi', 'bulunduğu_yer', 'bakım_tarihi', 'muayene_tarihi', 'kontrol_tarihi')
    search_fields = ('adi', 'bulunduğu_yer')
    list_filter = ('bulunduğu_yer',)

@admin.register(QRSorguKaydi)
class QRSorguKaydiAdmin(admin.ModelAdmin):
    list_display = ('ekipman', 'kullanici', 'ip_adresi', 'sorgulama_zamani')
    list_filter = ('ekipman', 'kullanici')
    search_fields = ('ekipman__adi', 'kullanici__username', 'ip_adresi')
    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "QR Sorgu Kayıtları"

        # Başlık satırı
        columns = ['Ekipman', 'Kullanıcı', 'IP Adresi', 'Sorgulama Zamanı']
        ws.append(columns)

        # Veriler
        for sorgu in queryset:
            ws.append([
                sorgu.ekipman.adi,
                sorgu.kullanici.username if sorgu.kullanici else 'Anonim',
                sorgu.ip_adresi,
                sorgu.sorgulama_zamani.strftime('%Y-%m-%d %H:%M:%S')
            ])

        # Kolon genişliklerini ayarla
        for col_num, column_title in enumerate(columns, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="qr_sorgu_kayitlari.xlsx"'
        wb.save(response)
        return response

    export_as_excel.short_description = "Seçili sorguları Excel olarak indir"

# GÖREV ATAMA İÇİN EKLEDİK
@admin.register(Gorev)
class GorevAdmin(admin.ModelAdmin):
    list_display = ('ekipman', 'teknisyen', 'durum', 'olusturma_tarihi')
    list_filter = ('durum', 'teknisyen')
    search_fields = ('ekipman__adi', 'teknisyen__username')
